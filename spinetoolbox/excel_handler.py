#############################################################################
# Copyright (C) 2017 - 2018 VTT Technical Research Centre of Finland
#
# This file is part of Spine Toolbox.
#
# Spine Toolbox is free software: you can redistribute it and/or modify
# it under the terms of the GNU Lesser General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
#############################################################################

"""
Module for MS Excel file handling.

:author: Pekka Savolainen <pekka.t.savolainen@vtt.fi>
:date:   21.8.2018
"""

import logging
from collections import Counter
from openpyxl import load_workbook


class ExcelHandler:
    """Class for reading (and an example how to write) Excel (.xlsx) files.

    Attributes:
        path (str): Absolute path to an Excel file
    """
    def __init__(self, path):
        """Class constructor."""
        self.path = path
        self.wb = None

    def load_wb(self):
        """Load Excel workbook."""
        logging.debug("Opening Excel Workbook {0}".format(self.path))
        try:
            self.wb = load_workbook(filename=self.path)
        except Exception:
            raise Exception

    def sheet_names(self):
        """Return a list of sheet names in the opened workbook."""
        return self.wb.get_sheet_names()

    def get_header_and_data(self, sheet_name):
        """Returns the header and data from the given sheet name. Checks that the header is complete.

        Args:
            sheet_name (str): Name of sheet with data

        Returns:
            Header in a dictionary and data in a list. header is an empty dictionary
             if it is not found, and data is None if it is not found.
        """
        sheet = self.wb.get_sheet_by_name(sheet_name)
        header = dict()
        n_rows = sheet.max_row
        n_columns = sheet.max_column
        logging.debug("Processing sheet {0}. Contains {1} rows and {2} columns".format(sheet, n_rows, n_columns))
        # Get all data on the sheet into a list
        data = list(sheet.rows)
        # Get header legend and header data rows
        try:
            header_legend_row = data.pop(0)
            header_row = data.pop(0)
        except IndexError:
            return [header, data]
        header_legend = [item.value for item in header_legend_row]  # legend row
        header_row = [item.value for item in header_row]  # [type, symbol name, filename, extension]
        # Scrap None from header_legend
        header_legend = [e for e in header_legend if e is not None]
        # Make header_legend items lower()
        header_legend = [item.lower() for item in header_legend]
        # Find item that contains 'type' in header legend
        type_item = [e for e in header_legend if 'type' in e]  # list
        if not type_item:
            logging.debug("Type column not found")
            header['type'] = ''
        else:
            _type = header_row[header_legend.index(type_item[0])]
            if not _type:
                header['type'] = ''
            else:
                header['type'] = _type.lower()
        # Find item that contains 'symbol' in header legend
        symbol_item = [e for e in header_legend if 'symbol' in e]  # list
        if not symbol_item:
            header['symbol'] = ''
        else:
            symbol = header_row[header_legend.index(symbol_item[0])]
            if not symbol:
                header['symbol'] = ''
            else:
                header['symbol'] = symbol
        # Find item that contains 'file' and 'name' in header legend
        filename_item = [e for e in header_legend if 'file' in e and 'name' in e]  # list
        if not filename_item:
            header['filename'] = ''
        else:
            filename = header_row[header_legend.index(filename_item[0])]
            if not filename:
                header['filename'] = ''
            else:
                header['filename'] = filename
        # Find item that contains 'extension'. NOTE: 'ext' not usable because this might be in cell 'filename (no ext.)'
        extension_item = [e for e in header_legend if 'extension' in e]  # list
        if not extension_item:
            header['extension'] = ''
        else:
            extension = header_row[header_legend.index(extension_item[0])]
            if not extension:
                header['extension'] = ''
            else:
                header['extension'] = extension
        # Return empty dict as header if all values in header are empty
        if not any(header.values()):
            return [dict(), data]
        # Return empty dict if only filename is found. Possibly old style data sheet.
        if not header['filename'] == '' and header['type'] == ''\
                and header['symbol'] == '' and header['extension'] == '':
            return [dict(), data]
        return [header, data]

    def process_set_data(self, data):
        """Process type set data.

        Args:
            data (list): List of tuples, where one tuple is one row.

        Returns:
            A Dictionary containing the lines to be written to a file.
        """
        d = dict()
        data = self.clean_data(data)  # Remove None rows and columns
        data.pop(0)  # Pop data header
        # Get all Setup names from first column and skip empty cells
        setup_list = [v[0].value for v in data if v[0].value is not None]
        setups = Counter(map(str.lower, setup_list)).keys()  # Collect Setups (lower-case) to a list
        for setup_name in setups:
            d[setup_name] = (list(), list())  # Initialize lists for each Setup
        # Iterate rows to make lines that are ready to be written into file
        for i in range(len(data)):
            row = [c.value for c in data[i]]
            setup = row.pop(0)
            if not setup:
                logging.error("No Setup on data (set) row {0}".format(i))
                continue
            if None in row or '' in row:  # Skip rows with missing data
                logging.error("Incomplete data (set) row {0}".format(i))
                continue
            d[setup.lower()][0].append(tuple(row))  # Append indices
            d[setup.lower()][1].append(None)  # TODO: Append explanatory text
        return d

    def process_parameter_data(self, data):
        """Process type parameter data.

        Args:
            data (list): List of tuples, where one tuple is one row.

        Returns:
            A Dictionary containing the lines to be written to a file.
        """
        d = dict()
        data = self.clean_data(data)  # Remove None rows and columns
        # for i in range(len(data)):
        #     row = [c.value for c in data[i]]
        #     logging.debug("Row {0}:{1}".format(i, row))
        data.pop(0)  # Pop data header
        # Get all Setup names from first column and skip empty cells
        setup_list = [v[0].value for v in data if v[0].value is not None]
        setups = Counter(map(str.lower, setup_list)).keys()  # Collect Setups (lower-case) to a list
        for setup_name in setups:
            d[setup_name] = (list(), list())  # Initialize lists for each Setup
        # Iterate rows to make lines that are ready to be written into file
        for i in range(len(data)):
            row = [c.value for c in data[i]]
            setup = row.pop(0)
            if not setup:
                logging.error("No Setup on data (parameter) row {0}".format(i))
                continue
            if None in row or '' in row:  # Skip rows with missing data
                logging.error("Missing values on data (parameter) row {0}".format(i))
                continue
            indices = tuple(row[:-1])
            try:
                value = float(row[-1])
            except ValueError:
                logging.error("No value on data (parameter) row {0}".format(i))
                continue
            d[setup.lower()][0].append(indices)  # Append indices
            d[setup.lower()][1].append(value)   # Append value
        return d

    def process_table_data(self, data):
        """Process type (pivot) table data.

        Args:
            data (list): List of tuples, where one tuple is one row.

        Returns:
            A Dictionary containing the lines to be written to a file.
        """
        d = dict()
        data = self.clean_data(data)  # Remove None rows and columns
        # Split data into middle_data and setup_set_value_data
        middle_data = list()
        setup_set_value_data = list()
        clean_middle_data = list()
        clean_setup_set_value_data = list()
        # NOTE: if Setup is missing, the row is appended to middle_data and then ignored.
        for i in range(len(data)):
            row = [c.value for c in data[i]]
            if row[0] is None:
                # This is Set2, Set3, ... data
                middle_data.append(row)
            else:
                # This is Setup, Set1 or Value data
                setup_set_value_data.append(row)
        # Clean NoneTypes and empty strings from data lists and convert items to string
        for m_row in middle_data:
            clean_m_row = [x for x in m_row if x is not None]
            clean_m_row = [str(x) for x in clean_m_row if x is not '']
            clean_middle_data.append(clean_m_row)
        for s_row in setup_set_value_data:
            clean_s_row = [x for x in s_row if x is not None]
            clean_s_row = [str(x) for x in clean_s_row if x is not '']
            clean_setup_set_value_data.append(clean_s_row)
        # Pop data header from setup_set_value_data
        clean_setup_set_value_data.pop(0)
        # Pop data header from clean_middle_data
        [x.pop(0) for x in clean_middle_data]
        # Get all Setup names from first column and skip empty cells
        setup_list = [v[0] for v in clean_setup_set_value_data if v[0] is not None]  # This probably never has Nones
        setups = Counter(map(str.lower, setup_list)).keys()  # Collect Setups (lower-case) to a list
        for setup_name in setups:
            d[setup_name] = (list(), list())  # Initialize lists for each Setup
        # Make written lines for each setup
        for j in range(len(clean_setup_set_value_data)):
            # j is row
            for i in range(len(clean_setup_set_value_data[0]) - 2):  # Subtract 2 (Setup and Set1 columns)
                # i is the index of values
                setup = clean_setup_set_value_data[j][0]  # NOTE: This is not None even if Setup is missing
                set1 = clean_setup_set_value_data[j][1]  # First set
                try:
                    sets = [x[i] for x in clean_middle_data]  # Sets 2, 3, ...
                except IndexError:
                    logging.error("IndexError in table data. No set in index {0}".format(i))
                    sets = list()
                try:
                    value = float(clean_setup_set_value_data[j][i+2])  # Value
                except IndexError:
                    logging.error("IndexError in table data. No value in index {0}".format(i+2))
                    value = None
                d[setup.lower()][0].append(tuple([set1] + sets))
                d[setup.lower()][1].append(value)
        return d

    # noinspection PyMethodMayBeStatic
    def clean_data(self, data):
        """Remove empty (None) rows and columns from the given list of tuples.

        Args:
            data (list): List of tuples with possible rows and columns with Nones.

        Returns:
            list of tuples with empty (None) rows and columns removed.
        """
        n_rows = len(data)
        n_columns = len(data[0])
        empty_rows = list()
        empty_columns = list()
        # Do rows
        for i in range(n_rows):
            row = [c.value for c in data[i]]
            if not any(row):
                # Empty row found
                empty_rows.insert(0, i)  # Insert to index 0
        # Remove empty rows
        for r in empty_rows:
            data.pop(r)
        # Transpose data
        data_t = list(zip(*data))
        # Do columns
        for j in range(n_columns):
            column = [c.value for c in data_t[j]]
            if not any(column):
                # Find empty columns
                empty_columns.insert(0, j)  # Insert to index 0
        # Remove empty columns
        for co in empty_columns:
            data_t.pop(co)
        # Transpose data again
        data = list(zip(*data_t))
        return data

    def read_data_sheet(self, sheet_name):
        """[OBSOLETE] Read data from the given sheet.

        Args:
            sheet_name (str): Name of sheet with data

        Returns:
            List of lists including filename, Setups and other data
        """
        sheet = self.wb.get_sheet_by_name(sheet_name)
        n_rows = sheet.max_row
        n_columns = sheet.max_column
        logging.debug("Processing {0}. Includes {1} rows and {2} columns".format(sheet, n_rows, n_columns))
        # Get all data on the sheet
        data = list(sheet.rows)
        # Get header row
        header = data.pop(0)
        if n_columns == 1 and n_rows == 1 and not header[0].value:
            logging.debug("No data found on {0}".format(sheet))
            return []
        if n_columns < 4:
            logging.error("Sheet should have at least 4 columns: Filename, Setup, Set, and Value")
            return []
        headers = [item.value for item in header]  # Append header values into list
        # Number of Set columns
        n_sets = n_columns - 3  # n_columns - filename column - setup column - value column
        sets = list()
        # All data
        filename = [v[0].value for v in data]
        setup = [v[1].value for v in data]
        # Get all sets into one list
        for i in range(n_sets):
            set_i = [v[i+2].value for v in data]
            sets.append(set_i)
        value = [v[-1].value for v in data]
        return [headers, filename, setup, sets, value, n_rows]

    # noinspection PyMethodMayBeStatic
    def make_data_lines(self, data):
        """[OBSOLETE] Make a dictionary where key is Setup
        name and value is a list containing the lines to be written.

        Args:
            data (list): Setups, sets, and values in a list of tuples
        """
        d = dict()
        sets = list()
        n_rows = len(data)
        # Number of Set columns
        n_sets = len(data[0]) - 2  # n_columns - setup column - value column
        setups = [v[0].value for v in data]  # Setup is in the first column
        for i in range(n_sets):  # Sets are between setup and value columns
            set_i = [v[i+1].value for v in data]
            sets.append(set_i)
        values = [v[-1].value for v in data]  # Value is in the last column
        # Make a dictionary key for each Setup
        ind_setups = Counter(setups).keys()
        for setup_name in ind_setups:
            d[setup_name] = list()
        # Collect data into dictionary
        for i in range(n_rows):
            if not sets[0][i]:  # Check that first set is not '' or None
                line = ''
            else:
                line = sets[0][i]
            if n_sets > 1:
                for j in range(n_sets-1):  # Subtract first set
                    if sets[j+1][i]:  # Skip '' and None sets
                        if line == '':  # This happens when first set is empty
                            line = sets[j+1][i]
                        else:
                            line += '.' + str(sets[j+1][i])  # Append new set on line
            # noinspection PyComparisonWithNone
            # Add '=' and value to line
            # Do not use if not value[i]: because if value is 0, it will not be written
            if values[i] == '' or values[i] == None:
                line += '=\n'
            else:
                line += '=' + str(values[i]) + '\n'
            d[setups[i]].append(line)
        return d

    # **** Example on how to write to an Excel file. ****

    # def export_to_excel(self):
    #     """Exports the selected data to the defined excel file.
    #     """
    #     python_date_format = '%Y-%m-%d %H:%M:%S'
    #     excel_date_format = 'yyyy-mm-dd hh:mm:ss'
    #     logging.debug('Exporting %s data' % self._title)
    #     # Check if directory exists and create if not
    #     file_dir = os.path.dirname(config.EXPORTED_DATA_FILE)
    #     if not os.path.exists(file_dir):
    #         logging.debug("Creating folder: %s" % file_dir)
    #         os.makedirs(file_dir)
    #
    #     # Open new file or load existing
    #     if os.path.exists(config.EXPORTED_DATA_FILE):
    #         logging.debug('Opening file %s' % config.EXPORTED_DATA_FILE)
    #         wbook = openpyxl.load_workbook(filename=config.EXPORTED_DATA_FILE)
    #     else:
    #         logging.debug('Creating file %s' % config.EXPORTED_DATA_FILE)
    #         wbook = openpyxl.Workbook()
    #
    #     # If the sheet already exists, delete it in order to replace it
    #     try:
    #         wbook.remove_sheet(wbook.get_sheet_by_name(self._title))
    #     except (KeyError, ValueError):
    #         pass
    #
    #     # Create sheet, write title, and export timestamp
    #     wsheet = wbook.create_sheet()
    #     wsheet.title = self._title
    #     wsheet.cell(None, 1, 1).value = self._title
    #     export_stamp = 'Exported: ' \
    #                    + datetime.now().strftime(python_date_format)
    #     wsheet.cell(None, 2, 1).value = export_stamp
    #
    #     sorted_scenarios = self.sort_scenarios()
    #     # Write the headers (configuration, scenario, date and label)
    #     i = 2
    #     for scenario in sorted_scenarios:
    #         conf_name = self.get_conf_name(scenario)
    #         wsheet.cell(None, 1, i).value = conf_name
    #         wsheet.cell(None, 2, i).value = scenario.scenario
    #         wsheet.cell(None, 3, i).value = scenario.created
    #         wsheet.cell(None, 3, i).number_format = excel_date_format
    #         wsheet.cell(None, 4, i).value = scenario.label
    #         i += 1
    #
    #     # Writes results for time plots and fill plots.
    #     if self._plot_type is PlotType.TIME_PLOT or self._plot_type is PlotType.FILL_PLOT:
    #         all_timestamps, output_lines = self.get_timestamps()
    #         i = 2
    #         # Writes the values for each scenario
    #         for scenario in sorted_scenarios:
    #             values = self.get_data_time_plots(scenario, all_timestamps)
    #             for j in range(len(values)):
    #                 wsheet.cell(None, j+5, i).value = values[j]
    #             i += 1
    #         # Writes the timestamps in the first row
    #         i = 5
    #         for line in output_lines:
    #             wsheet.cell(None, i, 1).value \
    #                 = datetime.strptime(line[:-1], python_date_format)
    #             wsheet.cell(None, i, 1).number_format = excel_date_format
    #             i += 1
    #
    #     # Delete default sheet and save file
    #     try:
    #         wbook.remove_sheet(wbook.get_sheet_by_name('Sheet'))
    #     except (KeyError, ValueError):
    #         pass
    #     # Handle error when file is being used
    #     if not self.save_and_close_excel():
    #         logging.debug("File not exported.")
    #
    # def save_and_close_excel(self):
    #     """ Handles the saving of the exported file
    #
    #     Args:
    #         wbook: workbook being written on.
    #
    #     Return:
    #         boolean flagging the success of the save.
    #     """
    #
    #     try:
    #         self.wb.save(self.path)
    #     except PermissionError:
    #         logging.exception("Permission error. File is probably reserved for another process.")
    #         return False
    #     return True
    #
