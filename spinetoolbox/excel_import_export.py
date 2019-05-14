######################################################################################################################
# Copyright (C) 2017 - 2018 Spine project consortium
# This file is part of Spine Toolbox.
# Spine Toolbox is free software: you can redistribute it and/or modify it under the terms of the GNU Lesser General
# Public License as published by the Free Software Foundation, either version 3 of the License, or (at your option)
# any later version. This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
# without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU Lesser General
# Public License for more details. You should have received a copy of the GNU Lesser General Public License along with
# this program. If not, see <http://www.gnu.org/licenses/>.
######################################################################################################################

"""
Functions to import and export from excel to spine database.

:author: P. Vennström (VTT)
:date:   21.8.2018
"""

# TODO: PEP8: Do not use bare except. Too broad exception clause

from collections import namedtuple
from itertools import groupby, islice, takewhile
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from spinedb_api import import_data
from operator import itemgetter


SheetData = namedtuple("SheetData", ["sheet_name", "class_name", "object_classes",
                                     "parameters", "parameter_values", "objects",
                                     "class_type"])


def import_xlsx_to_db(db, filepath):
    """reads excel file in 'filepath' and insert into database in mapping 'db'.
    Returns two list, one with succesful writes to database, one with errors
    when trying to write to database.

    Args:
        db (spinedb_api.DatabaseMapping): database mapping for database to write to
        filepath (str): str with filepath to excel file to read from

    Returns:
        (Int, List) Returns number of inserted items and a list of
        error information on all failed writes
    """

    obj_data, rel_data, error_log = read_spine_xlsx(filepath)

    object_classes = []
    objects = []
    object_parameters = []
    object_values = []
    for sheet in obj_data:
        object_classes.append(sheet.class_name)
        objects.extend([(sheet.class_name, o) for o in sheet.objects])
        object_parameters.extend([(sheet.class_name, o) for o in sheet.parameters])
        d_getter = itemgetter(*[1,2,0,3])
        object_values.extend([(sheet.class_name,) + d_getter(d) for d in sheet.parameter_values])

    rel_classes = []
    rels = []
    rel_parameters = []
    rel_values = []
    for sheet in rel_data:
        num_oc = len(sheet.object_classes)
        rel_getter = itemgetter(*range(1,num_oc + 1 ))
        d_getter = itemgetter(*[num_oc+1, 0, num_oc+2])
        rel_classes.append((sheet.class_name, sheet.object_classes))
        rels.extend([(sheet.class_name, o) for o in sheet.objects])
        rel_parameters.extend([(sheet.class_name, o) for o in sheet.parameters])
        rel_values.extend([(sheet.class_name, rel_getter(d)) + d_getter(d) for d in sheet.parameter_values])
    
    
    object_values = [o[:-2] + (o[-1],) for o in object_values]
    rel_values = [rel[:-2] + (rel[-1],) for rel in rel_values]

    num_imported, errors = import_data(db, object_classes, rel_classes, object_parameters, rel_parameters, objects, rels, object_values, rel_values)
    error_log.extend(errors)
    return num_imported, error_log


def get_objects_and_parameters(db):
    """Exports all object data from spine database into unstacked list of lists

    Args:
        db (spinedb_api.DatabaseMapping): database mapping for database

    Returns:
        (List, List) First list contains parameter data, second one json data
    """

    # get all objects
    obj = db.object_list().add_columns(db.ObjectClass.name.label('class_name')).\
        filter(db.ObjectClass.id == db.Object.class_id).all()

    # get all object classes
    obj_class = db.object_class_list().all()

    # get all parameter values
    pval = db.object_parameter_value_list().all()

    # get all parameter definitions
    par = db.object_parameter_definition_list().all()

    # make all in same format
    par = [(p.object_class_name, None, p.parameter_name, None) for p in par]
    pval = [(p.object_class_name, p.object_name, p.parameter_name, json.loads(p.value)) for p in pval]
    obj = [(p.class_name, p.name, None, None) for p in obj]
    obj_class = [(p.name, None, None, None) for p in obj_class]

    object_and_par = pval + par + obj + obj_class
    
    object_par = []
    object_json = []
    for d in object_and_par:
        if isinstance(d[3],list):
            object_json.append(d)
            object_par.append(d[:-1] + (None,))
        else:
            if isinstance(d[3], dict):
                d = d[:-1] + (json.dumps(d[3]),)
            elif isinstance(d[3], str):
                d = d[:-1] + ('"' + d[3] + '"',)
            object_par.append(d)


    return object_par, object_json


def get_relationships_and_parameters(db):
    """Exports all relationship data from spine database into unstacked list of lists

    Args:
        db (spinedb_api.DatabaseMapping): database mapping for database

    Returns:
        (List, List) First list contains parameter data, second one json data
    """

    rel_class = db.wide_relationship_class_list().all()
    rel = db.wide_relationship_list().all()
    rel_par = db.relationship_parameter_definition_list().all()
    rel_par_value = db.relationship_parameter_value_list().all()

    rel_class_id_2_name = {rc.id: rc.name for rc in rel_class}

    out_data = [[r.relationship_class_name,
                 r.object_name_list,
                 r.parameter_name,
                 json.loads(r.value)] for r in rel_par_value]

    rel_with_par = set(r.object_name_list for r in rel_par_value)
    rel_without_par = [[rel_class_id_2_name[r.class_id], r.object_name_list, None, None]
                       for r in rel if r.object_name_list not in rel_with_par]

    rel_class_par = [[r.relationship_class_name, None, r.parameter_name, None]
                     for r in rel_par]

    rel_class_with_par = [r.relationship_class_name for r in rel_par]
    rel_class_without_par = [[r.name, None, None, None]
                             for r in rel_class if r.name not in rel_class_with_par]

    rel_data = out_data + rel_without_par + rel_class_par + rel_class_without_par

    rel_par = []
    rel_json = []
    for d in rel_data:
        if isinstance(d[3],list):
            rel_json.append(d)
            rel_par.append(d[:-1] + [None])
        else:
            if isinstance(d[3], dict):
                d[3] = json.dumps(d[3])
            elif isinstance(d[3], str):
                d = d[:-1] + ['"' + d[3] + '"',]
            rel_par.append(d)

    return rel_par, rel_json, rel_class


def unstack_list_of_tuples(data, headers, key_cols, value_name_col, value_col):
    """Unstacks list of lists or list of tuples and creates a list of namedtuples
    whit unstacked data (pivoted data)

    Args:
        data (List[List]): List of lists with data to unstack
        headers (List[str]): List of header names for data
        key_cols (List[Int]): List of index for column that are keys, columns to not unstack
        value_name_col (Int): index to column containing name of data to unstack
        value_col (Int): index to column containing value to value_name_col

    Returns:
        (List[List]): List of list with headers in headers list
        (List): List of header names for each item in inner list
    """
    # find header names
    if isinstance(value_name_col, list) and len(value_name_col) > 1:
        value_name_getter = itemgetter(*value_name_col)
        value_names = sorted(set(value_name_getter(x) for x in data if not any(i is None for i in value_name_getter(x))))
    else:
        if isinstance(value_name_col, list):
            value_name_col = value_name_col[0]
        value_name_getter = itemgetter(value_name_col)
        value_names = sorted(set(x[value_name_col] for x in data if x[value_name_col] is not None))



    key_names = [headers[n] for n in key_cols]
    #value_names = sorted(set(x[value_name_col] for x in data if x[value_name_col] is not None))
    headers = key_names + value_names

    # remove data with invalid key cols
    keyfunc = lambda x: [x[k] for k in key_cols]
    data = [x for x in data if None not in keyfunc(x)]
    data = sorted(data, key=keyfunc)

    # unstack data
    data_list_out = []
    for k, k_data in groupby(data, key=keyfunc):
        if None in k:
            continue
        line_data = [None]*len(value_names)
        for d in k_data:
            if value_name_getter(d) in value_names and d[value_col] is not None:
                line_data[value_names.index(value_name_getter(d))] = d[value_col]
        data_list_out.append(k + line_data)

    return data_list_out, headers


def stack_list_of_tuples(data, headers, key_cols, value_cols):
    """Stacks list of lists or list of tuples and creates a list of namedtuples
    with stacked data (unpivoted data)

    Args:
        data (List[List]): List of lists with data to unstack
        headers (List[str]): List of header names for data
        key_cols (List[Int]): List of index for columns that are keys
        value_cols (List[Int]): List of index for columns containing values to stack

    Returns:
        (List[namedtuple]): List of namedtuples whit fields given by headers
        and 'parameter' and 'value' which contains stacked values
    """
    value_names = [headers[n] for n in value_cols]
    key_names = [headers[n] for n in key_cols]
    new_tuple_names = key_names + ["parameter", "value"]
    NewDataTuple = namedtuple("Data", new_tuple_names)
    # takes unstacked data and duplicates columns in key_cols and then zips
    # them with values in value_cols
    new_data_list = [list(map(
            NewDataTuple._make,
            [a+[b]+[c] for a, b, c in zip([[dl[k] for k in key_cols]]*len(value_cols),
                                          value_names, [dl[vk] for vk in value_cols])])) for dl in data]
    new_data_list = [item for sublist in new_data_list for item in sublist]
    return new_data_list


def unpack_json_parameters(data, json_index):
    out_data = []
    for data_row in data:
        json_data = json.loads(data_row[json_index].replace("\n", ""))

        if json_index == 0:
            key_cols = [list(data_row[json_index+1:])]*len(json_data)
        else:
            key_cols = [list(data_row[:json_index]) +
                        list(data_row[json_index+1:])]*len(json_data)

        out_data += [a + [b] + [c] for a, b, c in
                     zip(key_cols, range(0, len(json_data), 1), json_data)]
    return out_data


def pack_json_parameters(data, key_cols, value_col, index_col=None):
    out_data = []
    # group by keys cols
    keyfunc = lambda x: [x[k] for k in key_cols]
    data = sorted(data, key=keyfunc)
    for key, grouped in groupby(data, key=keyfunc):
        # sort if index is given.
        if index_col is not None:
            grouped = sorted(grouped, key=lambda x: x[index_col])

        # pack values into json
        values = [g[value_col] for g in grouped]
        json_val = json.dumps(values)
        out_data.append(key + [json_val])
    return out_data


def get_unstacked_relationships(db):
    """Gets all data for relationships in a unstacked list of list

    Args:
        db (spinedb_api.DatabaseMapping): database mapping for database

    Returns:
        (List, List): Two list of data for relationship, one with parameter values
        and the second one with json values
    """
    data, data_json, rel_class = get_relationships_and_parameters(db)

    class_2_obj_list = {rc.name: rc.object_class_name_list.split(',') for rc in rel_class}

    keyfunc = lambda x: x[0]
    data_json = sorted(data_json, key=keyfunc)
    parsed_json = []
    # json data, split by relationship class
    for k, v in groupby(data_json, key=keyfunc):
        json_vals = []
        for row in v:
            rel_list = row[1].split(',')
            parameter = row[2]
            val = row[3]
            json_vals.append([rel_list+[parameter], val])
        if json_vals:
            object_classes = class_2_obj_list[k]
            parsed_json.append([k, object_classes, json_vals])

    # parameter data, split by relationship class
    stacked_rels = []
    data = sorted(data, key=keyfunc)
    for k, v in groupby(data, key=keyfunc):
        values = list(v)
        rel, par_names = unstack_list_of_tuples(values, ["relationship_class", "relationship", "parameter", "value"], [0, 1], 2, 3)
        if rel:
            parameters = par_names[2:]
        else:
            parameters = list(set([p[2] for p in values]))
        rel = [r[1].split(',') + list(r[2:]) for r in rel]
        object_classes = class_2_obj_list[k]
        stacked_rels.append([k, rel, object_classes, parameters])
    return stacked_rels, parsed_json


def get_unstacked_objects(db):
    """Gets all data for objects in a unstacked list of list

    Args:
        db (spinedb_api.DatabaseMapping): database mapping for database

    Returns:
        (List, List): Two list of data for objects, one with parameter values
        and the second one with json values
    """
    data, data_json = get_objects_and_parameters(db)

    keyfunc = lambda x: x[0]

    parsed_json = []
    data_json = sorted(data_json, key=keyfunc)
    for k, v in groupby(data_json, key=keyfunc):
        json_vals = []
        for row in v:
            obj = row[1]
            parameter = row[2]
            val = row[3]
            json_vals.append([[obj, parameter], val])
        if json_vals:
            parsed_json.append([k, [k], json_vals])

    stacked_obj = []
    data = sorted(data, key=keyfunc)
    for k, v in groupby(data, key=keyfunc):
        values = list(v)
        obj, par_names = unstack_list_of_tuples(values, ["object_class", "object", "parameter", "value"], [0, 1], 2, 3)
        if obj:
            parameters = par_names[2:]
        else:
            parameters = list(set([p[2] for p in values if p[2] is not None]))
        obj = [[o[1]] + list(o[2:]) for o in obj]
        object_classes = [k]
        stacked_obj.append([k, obj, object_classes, parameters])
    return stacked_obj, parsed_json


def write_relationships_to_xlsx(wb, relationship_data):
    """Writes Classes, parameter and parameter values for relationships.
    Writes one sheet per relationship class.

    Args:
        wb (openpyxl.Workbook): excel workbook to write too.
        relationship_data (List[List]): List of lists containing relationship
        data give by function get_unstacked_relationships
    """
    for rel in relationship_data:
        ws = wb.create_sheet()

        # try setting the sheetname to relationship class name
        # sheet name can only be 31 chars log
        title = "rel_" + rel[0]
        if len(title) < 32:
            ws.title = title

        ws['A1'] = "Sheet type"
        ws['A2'] = "relationship"
        ws['B1'] = "Data type"
        ws['B2'] = "Parameter"
        ws['C1'] = "relationship class name"
        ws['C2'] = rel[0]
        ws['D1'] = "Number of relationship dimensions"
        ws['D2'] = len(rel[2])
        ws['E1'] = "Number of pivoted relationship dimensions"
        ws['E2'] = 0

        for c, val in enumerate(rel[2]):
            ws.cell(row=4, column=c+1).value = val

        for c, val in enumerate(rel[3]):
            ws.cell(row=4, column=len(rel[2]) + 1 + c).value = val

        start_row = 5
        start_col = 1
        for r, line in enumerate(rel[1]):
            for c, val in enumerate(line):
                ws.cell(row=start_row + r, column=start_col + c).value = val


def write_json_array_to_xlsx(wb, data, sheet_type):
    """Writes json array data for object classes and relationship classes.
    Writes one sheet per relationship/object class.

    Args:
        wb (openpyxl.Workbook): excel workbook to write too.
        data (List[List]): List of lists containing json data give by function
        get_unstacked_objects and get_unstacked_relationships
        sheet_type (str): str with value "relationship" or "object" telling if data is for a relationship or object
    """
    for i, d in enumerate(data):
        if sheet_type == "relationship":
            sheet_title = "json_"
        elif sheet_type == "object":
            sheet_title = "json_"
        else:
            raise ValueError("sheet_type must be a str with value 'relationship' or 'object'")

        ws = wb.create_sheet()
        # sheet name can only be 31 chars log
        title = sheet_title + d[0]
        if len(title) < 32:
            ws.title = title
        else:
            ws.title = '{}_json{}'.format(sheet_type, i)

        ws['A1'] = "Sheet type"
        ws['A2'] = sheet_type
        ws['B1'] = "Data type"
        ws['B2'] = "json array"
        ws['C1'] = sheet_type + " class name"
        ws['C2'] = d[0]

        if sheet_type == "relationship":
            ws['D1'] = "Number of relationship dimensions"
            ws['D2'] = len(d[1])

        title_rows = d[1]+["json parameter"]
        for c, val in enumerate(title_rows):
            ws.cell(row=4+c, column=1).value = val

        start_row = 4 + len(title_rows)
        for col, obj_list in enumerate(d[2]):
            for obj_iter, obj in enumerate(obj_list[0]):
                ws.cell(row=4 + obj_iter, column=2 + col).value = obj

            for row_iter, json_val in enumerate(obj_list[1]):
                ws.cell(row=start_row + row_iter, column=2 + col).value = json_val


def write_objects_to_xlsx(wb, object_data):
    """Writes Classes, parameter and parameter values for objects.
    Writes one sheet per relationship/object class.

    Args:
        wb (openpyxl.Workbook): excel workbook to write too.
        object_data (List[List]): List of lists containing relationship data give by function get_unstacked_objects
    """

    for i, obj in enumerate(object_data):
        ws = wb.create_sheet()

        # try setting the sheetname to object class name
        # sheet name can only be 31 chars log
        title = "obj_" + obj[0]
        if len(title) < 32:
            ws.title = title
        else:
            ws.title = "object_class{}".format(i)

        ws['A1'] = "Sheet type"
        ws['A2'] = "object"
        ws['B1'] = "Data type"
        ws['B2'] = "Parameter"
        ws['C1'] = "object class name"
        ws['C2'] = obj[0]

        for c, val in enumerate(obj[2]):
            ws.cell(row=4, column=c+1).value = val

        for c, val in enumerate(obj[3]):
            ws.cell(row=4, column=len(obj[2]) + 1 + c).value = val

        start_row = 5
        start_col = 1
        for r, line in enumerate(obj[1]):
            for c, val in enumerate(line):
                ws.cell(row=start_row + r, column=start_col + c).value = val


def export_spine_database_to_xlsx(db, filepath):
    """Writes all data in a spine database into an excel file.

    Args:
        db (spinedb_api.DatabaseMapping): database mapping for database.
        filepath (str): str with filepath to save excel file to.
    """
    obj_data, obj_json_data = get_unstacked_objects(db)
    rel_data, rel_json_data = get_unstacked_relationships(db)
    wb = Workbook()
    write_relationships_to_xlsx(wb, rel_data)
    write_objects_to_xlsx(wb, obj_data)
    write_json_array_to_xlsx(wb, obj_json_data, "object")
    write_json_array_to_xlsx(wb, rel_json_data, "relationship")
    wb.save(filepath)
    wb.close()


def read_spine_xlsx(filepath):
    """reads all data from a excel file where the sheets are in valid spine data format

    Args:
        filepath (str): str with filepath to excel file to read from.
    """

    #

    # read_only=true doesn't seem to close the file properly, possible solution if
    # speed is needed is to do following:
    # with open(xlsx_filename, "rb") as f:
    #   in_mem_file = io.BytesIO(f.read())
    #       wb = load_workbook(in_mem_file, read_only=True)
    wb = load_workbook(filepath, read_only=False)
    sheets = wb.sheetnames
    ErrorLogMsg = namedtuple('ErrorLogMsg',('msg','db_type','imported_from','other'))

    obj_data = []
    rel_data = []
    obj_json_data = []
    rel_json_data = []
    error_log = []

    # read all sheets
    for s in sheets:
        ws = wb[s]

        # check if valid
        if not validate_sheet(ws):
            continue

        sheet_type = ws['A2'].value.lower()
        sheet_data = ws['B2'].value.lower()

        if sheet_data == "parameter":
            # read sheet with data type: 'parameter'
            try:
                data = read_parameter_sheet(ws)
                if sheet_type == "relationship":
                    rel_data.append(data)
                else:
                    obj_data.append(data)
            except Exception as e:
                error_log.append(ErrorLogMsg("Error reading sheet {}: {}".format(ws.title, e), "sheet", filepath, ''))
        elif sheet_data == "json array":
            # read sheet with data type: 'json array'
            try:
                data = read_json_sheet(ws, sheet_type)
                if sheet_type == "relationship":
                    rel_json_data.append(data)
                else:
                    obj_json_data.append(data)
            except Exception as e:
                error_log.append(ErrorLogMsg("Error reading sheet {}: {}".format(ws.title, e), "sheet", filepath, ''))
    wb.close()

    # merge sheets that have the same class.
    obj_data, el = merge_spine_xlsx_data(obj_data + obj_json_data)
    error_log = error_log + el

    rel_data, el = merge_spine_xlsx_data(rel_data + rel_json_data)
    error_log = error_log + el

    return obj_data, rel_data, error_log


def merge_spine_xlsx_data(data):
    """Merge data from different sheets with same object class or
    relationship class.

    Args:
        data (List(SheetData)): list of SheetData

    Returns:
        (List[SheetData]): List of SheetData with only one relationship/object class per item
    """
    error_log = []
    new_data = []
    data = sorted(data, key=lambda x: x.class_name)
    for class_name, values in groupby(data, key=lambda x: x.class_name):
        values = list(values)
        if len(values) == 1:
            # only one sheet
            new_data.append(values[0])
            continue
        else:
            # if more than one SheetData with same class_name
            sheet_name = values[0].sheet_name
            object_classes = values[0].object_classes
            parameters = values[0].parameters
            parameter_values = values[0].parameter_values
            objects = values[0].objects
            class_type = values[0].class_type

            # skip first sheet
            iter_values = iter(values)
            next(iter_values)
            for v in iter_values:
                # make sure that the new sheet has same object_classes that first
                if v.object_classes != object_classes:
                    error_log.append(["sheet", v.sheet_name, "sheet {} as different "
                                                             "object_classes than sheet {} for class {}"
                                     .format(v.sheet_name, sheet_name, class_name)])
                    continue
                parameters = parameters + v.parameters
                objects = objects + v.objects
                parameter_values = parameter_values + v.parameter_values

            # make unique again
            parameters = list(set(parameters))

            if len(object_classes) > 1:
                keyfunc = lambda x: [x[i] for i,_ in enumerate(object_classes)]
                objects = sorted(objects, key=keyfunc)
                objects = list(k for k, _ in groupby(objects, key=keyfunc))
            else:
                objects = list(set(objects))

            new_data.append(SheetData(sheet_name=sheet_name,
                                      class_name=class_name,
                                      object_classes=object_classes,
                                      parameters=parameters,
                                      parameter_values=parameter_values,
                                      objects=objects,
                                      class_type=class_type))

    return new_data, error_log


def validate_sheet(ws):
    """Checks if supplied sheet is a valid import sheet for spine.

    Args:
        ws (openpyxl.workbook.worksheet): worksheet to validate

    Returns:
        (bool): True if sheet is valid, False otherwise
    """
    sheet_type = ws['A2'].value
    sheet_data = ws['B2'].value

    if not isinstance(sheet_type, str):
        return False
    if not isinstance(sheet_data, str):
        return False
    if sheet_type.lower() not in ["relationship", "object"]:
        return False
    if sheet_data.lower() not in ["parameter", "json array"]:
        return False

    if sheet_type.lower() == "relationship":
        rel_dimension = ws['D2'].value
        rel_name = ws['C2'].value
        if not isinstance(rel_name, str):
            return False
        if not rel_name:
            return False
        if not isinstance(rel_dimension, int):
            return False
        if not rel_dimension > 1:
            return False
        if sheet_data.lower() == 'parameter':
            rel_row = read_2d(ws, 4, 4, 1, rel_dimension)[0]
        else:
            rel_row = read_2d(ws, 4, 4 + rel_dimension - 1, 1, 1)
            rel_row = [r[0] for r in rel_row]
        if None in rel_row:
            return False
        if not all(isinstance(r, str) for r in rel_row):
            return False
        if not all(r for r in rel_row):
            return False
    elif sheet_type.lower() == "object":
        obj_name = ws['C2'].value
        if not isinstance(obj_name, str):
            return False
        if not obj_name:
            return False
    else:
        return False
    return True


def read_json_sheet(ws, sheet_type):
    """Reads a sheet containg json array data for objects and relationships

    Args:
        ws (openpyxl.workbook.worksheet): worksheet to read from
        sheet_type (str): str with value "relationship" or "object" telling if sheet is a relationship or object sheet

    Returns:
        (List[SheetData])
    """
    if sheet_type == "relationship":
        dim = ws['D2'].value
    else:
        dim = 1

    path = ["object" + str(i) for i in range(dim)]

    class_name = ws['C2'].value

    object_classes = []
    for i in range(4, 4 + dim, 1):
        object_classes.append(ws["A" + str(i)].value)

    # search row for until first empty cell
    add_if_not_break = 1
    for c, cell in enumerate(ws[4]):
        if c > 0:
            if cell.value is None:
                add_if_not_break = 0
                break
    read_cols = range(1, c + add_if_not_break)

    json_data = []
    # red columnwise from second column.
    rows = ws.iter_rows()
    obj_path = []
    parameters = []
    for r, row in enumerate(rows):
        if 2 < r < 3+dim:
            # get object path
            obj_path.append([cell.value for i, cell in enumerate(row) if i in read_cols])
        elif r == 3+dim:
            # get parameter name
            parameters = [cell.value for i, cell in enumerate(row) if i in read_cols]
            break

    data = [[cell.value for i, cell in enumerate(row) if i in read_cols] for row in rows]

    # pivot data
    obj_path = [[obj_path[r][c] for r in range(len(obj_path))] for c in range(len(obj_path[0]))]
    data = [[data[r][c] for r in range(len(data))] for c in range(len(data[0]))]


    Data = namedtuple("Data", ["parameter_type"] + path + ["parameter", "value"])
    if data:
        for objects, parameter, data_list in zip(obj_path, parameters, data):
            # save values if there is json data, a parameter name
            # and the obj_path doesn't contain None.
            packed_json = json.dumps(list(takewhile(lambda x: x is not None, data_list)))
            json_data.append(Data._make(["json"] + objects + [parameter, packed_json]))

    return SheetData(sheet_name=ws.title,
                     class_name=class_name,
                     object_classes=object_classes,
                     parameters=list(set(parameters)),
                     parameter_values=json_data,
                     objects=[],
                     class_type=sheet_type)


def read_parameter_sheet(ws):
    """Reads a sheet containg parameter data for objects and relationships

    Args:
        ws (openpyxl.workbook.worksheet): worksheet to read from

    Returns:
        (List[SheetData])
    """
    sheet_type = ws['A2'].value.lower()
    class_name = ws['C2'].value

    if sheet_type == "object":
        dim = 1
    elif sheet_type == "relationship":
        dim = ws['D2'].value
    else:
        raise ValueError("sheet_type must be a str with value 'relationship' or 'object'")

    # object classes
    object_classes = read_2d(ws, 4, 4, 1, dim)[0]

    # read all columns to the right of the number of cells in dim. Read until
    # encounters a empty cell.
    parameters = []
    read_cols = []
    for c, cell in enumerate(ws[4]):
        if cell.value is None:
            break
        elif c >= dim:
            parameters.append(cell.value)
    read_cols = range(0, c + 1)

    # get data
    rows = islice(ws.iter_rows(), 4, None)
    try:
        data = [[cell.value for i, cell in enumerate(row) if i in read_cols] for row in rows]
    except StopIteration:
        data = []

    keyfunc = lambda x: [x[i] for i, _ in enumerate(object_classes)]

    # remove data where not all dimensions exists
    data = [d for d in data if not None in keyfunc(d)]

    data_parameter = []
    if parameters:
        # add that parameter type type should be "value"
        data = [["value"] + d for d in data]
        keyfunc = lambda x: [x[i+1] for i, _ in enumerate(object_classes)]
        key_cols = list(range(0, dim+1, 1))
        val_cols = list(range(dim+1, dim+len(parameters)+1))
        headers = ["parameter_type"] + ["object" + str(x) for x in range(dim)] + parameters
        data_parameter = stack_list_of_tuples(data, headers, key_cols, val_cols)
        data_parameter = [d for d in data_parameter if d.value is not None]

    # find unique relationships from data
    data = sorted(data, key=keyfunc)
    objects = list(k for k, _ in groupby(data, key=keyfunc))
    if dim == 1:
        # flatten list if only one object per row
        objects = [item for sublist in objects for item in sublist]

    return SheetData(sheet_name=ws.title,
                     class_name=class_name,
                     object_classes=object_classes,
                     parameters=parameters,
                     parameter_values=data_parameter,
                     objects=objects,
                     class_type=sheet_type)


def read_2d(ws, start_row=1, end_row=1, start_col=1, end_col=1):
    """Reads a 2d area from worksheet into a list of lists where each line is
    the inner list.

    Args:
        ws (openpyxl.workbook.worksheet): Worksheet to look in
        start_row (Integer): start row to read, 1-indexed (as excel)
        end_row (Integer): row to read to, 1-indexed (as excel)
        start_col (Integer): start column to read, 1-indexed (as excel)
        end_col (Integer): end column to read to, 1-indexed (as excel)

    Returns:
        (List) List of all lines read.
    """
    end_col = get_column_letter(end_col)
    start_col = get_column_letter(start_col)
    xl_index = '{}{}:{}{}'.format(start_col, start_row, end_col, end_row)
    values = [[inner.value for inner in outer] for outer in ws[xl_index]]
    return values


def max_col_in_row(ws, row=1):
    """Finds max col index for given row. If no data exists on row, returns 1.

    Args:
        ws (openpyxl.workbook.worksheet): Worksheet to look in
        row (Integer): index for row to search, 1-indexed (as excel)

    Returns:
        (Integer) column index of last cell with value.
    """
    # TODO: Does this work if ws[row] is empty? In that case 'cell' is not initialized.
    for cell in reversed(ws[row]):
        if cell.value is not None:
            break
    return cell.col_idx
