# -*- coding: utf-8 -*-

from operator import itemgetter
from itertools import islice

from io_api import FileImportTemplate

from PySide2.QtWidgets import QWidget, QFormLayout, QLabel, QLineEdit, QCheckBox, QSpinBox, QGroupBox, QVBoxLayout
from PySide2.QtCore import QObject, Signal

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

class ExcelConnector(FileImportTemplate):
    def __init__(self):
        super(ExcelConnector, self).__init__()

        self._wb = None
        self.sheet_options = {}
        self._option_widget = ExcelOptionWidget()
        self.current_sheet = None
        
        self._option_widget.optionsChanged.connect(self._new_options)

    def _new_options(self):
        self.refreshDataRequest.emit()
    
    def set_table(self, sheet):
        if self.current_sheet:
            options = {"header": self._option_widget.first_row_as_header,
                       "row": self._option_widget.skip_rows,
                       "column": self._option_widget.skip_columns,
                       "read_until_col": self._option_widget.read_until_col,
                       "read_until_row": self._option_widget.read_until_row}
            self.sheet_options.update({self.current_sheet: options})
        self.current_sheet = sheet
        if sheet in self.sheet_options:
            self._option_widget.update_values(**self.sheet_options[sheet])
        else:
            self._option_widget.update_values()
    
    @property
    def file_filter(self):
        """
        Return filter string for file, change return for specific filter.
        """
        return '*.xlsx;*.xlsm;*.xltx;*.xltm'
    
    @property
    def tables(self):
        return self._wb.sheetnames
    
    @property
    def source_name(self):
        """Name of data source, must return string"""
        return 'Excel'
    
    @property
    def can_have_multiple_tables(self):
        return True
        
    def source_selector(self):
        filename, action = self.select_file()
        if not filename:
            return False
        self._wb = load_workbook(filename, read_only=True)
        return True
    
    def preview_data(self, table):
        return self.read_data(table, max_rows=100)
    
    def read_data(self, table, max_rows=None):
        """
        Return data read from data source table in table. If max_rows is 
        specified only that number of rows.
        """
        if not self._wb:
            return [], []
        
        if max_rows != None and type(max_rows) != int:
            raise TypeError('max_rows must be int')

        if not table in self._wb:
            # table not found
            return [], []
        ws = self._wb[table]
        
        has_header = self._option_widget.first_row_as_header
        skip_rows = self._option_widget.skip_rows
        skip_columns = self._option_widget.skip_columns
        stop_at_empty_col = self._option_widget.read_until_col
        stop_at_empty_row = self._option_widget.read_until_row

        
        
        # find first empty col in top row and use that as a stop
        read_to_col = None
        if stop_at_empty_col:
            try:
                first_row = next(islice(ws.iter_rows(), skip_rows, None))
                for i, c in enumerate(islice(first_row,skip_columns, None)):
                    if c.value is None:
                        read_to_col = i + skip_columns
                        break
            except StopIteration:
                # no data
                pass
        
        if any(r is None for r in [max_rows,skip_rows]):
            end_iter = None
        else:
            end_iter = max_rows+skip_rows
        data = []
        header = []
        rows = ws.iter_rows()
        rows = islice(rows, skip_rows, end_iter)
        # find header if it has one
        if has_header:
            try:
                header = [c.value for c in islice(next(rows),skip_columns, read_to_col)]
            except StopIteration:
                # no data
                pass

        data = []
        if stop_at_empty_row:
            for row in rows:
                row_data = [c.value for c in islice(row, skip_columns, read_to_col)]
                if row_data and row_data[0] is None:
                    break
                data.append(row_data)
        else:
            for row in rows:
                data.append([c.value for c in islice(row, skip_columns, read_to_col)])
        
        return data, header
    
    def option_widget(self):
        """
        Return a Qwidget with options for reading data from a table in source
        """
        return self._option_widget

class ExcelOptionWidget(QWidget):
    optionsChanged = Signal()
    def __init__(self, parent=None):
        super().__init__(parent)
        
        # state
        self.block_signals = False
        self.first_row_as_header = False
        self.skip_rows = 0
        self.skip_columns = 0
        self.read_until_row = False
        self.read_until_col = False
        
        # ui
        self._ui_skip_row = QSpinBox()
        self._ui_skip_col = QSpinBox()
        self._ui_header = QCheckBox()
        self._ui_read_until_col = QCheckBox()
        self._ui_read_until_row = QCheckBox()
        
        self._ui_skip_row.setMinimum(0)
        self._ui_skip_col.setMinimum(0)
        
        # layout
        groupbox = QGroupBox("Excel options")
        self.setLayout(QVBoxLayout())
        layout = QFormLayout()
        layout.addRow(QLabel("Skip rows:"), self._ui_skip_row)
        layout.addRow(QLabel("Skip columns:"), self._ui_skip_col)
        layout.addRow(QLabel("Header in first row:"), self._ui_header)
        layout.addRow(QLabel("Read until first empty column:"), self._ui_read_until_col)
        layout.addRow(QLabel("Read until first empty row:"), self._ui_read_until_row)
        groupbox.setLayout(layout)
        self.layout().addWidget(groupbox)
        
        # connect signals
        self._ui_skip_row.valueChanged.connect(self._skip_row_change)
        self._ui_skip_col.valueChanged.connect(self._skip_col_change)
        self._ui_header.stateChanged.connect(self._header_change)
        self._ui_read_until_col.stateChanged.connect(self._read_until_col_change)
        self._ui_read_until_row.stateChanged.connect(self._read_until_row_change)
        
    def update_values(self, header=False, row=0, column=0, read_until_row=False, read_until_col=False):
        self.block_signals = True
        self._ui_skip_row.setValue(row)
        self._ui_skip_col.setValue(column)
        self._ui_header.setChecked(header)
        self._ui_read_until_row.setChecked(read_until_row)
        self._ui_read_until_col.setChecked(read_until_col)
        self.block_signals = False
        
    def _header_change(self, new_bool):
        self.first_row_as_header = new_bool
        if not self.block_signals:
            self.optionsChanged.emit()
    
    def _read_until_row_change(self, new_bool):
        self.read_until_row = new_bool
        if not self.block_signals:
            self.optionsChanged.emit()
    
    def _read_until_col_change(self, new_bool):
        self.read_until_col = new_bool
        if not self.block_signals:
            self.optionsChanged.emit()
    
    def _skip_row_change(self, new_num):
        self.skip_rows = new_num
        if not self.block_signals:
            self.optionsChanged.emit()
    
    def _skip_col_change(self, new_num):
        self.skip_columns = new_num
        if not self.block_signals:
            self.optionsChanged.emit()
            
        
        

