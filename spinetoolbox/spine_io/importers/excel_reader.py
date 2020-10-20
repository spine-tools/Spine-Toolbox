######################################################################################################################
# Copyright (C) 2017-2020 Spine project consortium
# This file is part of Spine Toolbox.
# Spine Toolbox is free software: you can redistribute it and/or modify it under the terms of the GNU Lesser General
# Public License as published by the Free Software Foundation, either version 3 of the License, or (at your option)
# any later version. This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
# without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU Lesser General
# Public License for more details. You should have received a copy of the GNU Lesser General Public License along with
# this program. If not, see <http://www.gnu.org/licenses/>.
######################################################################################################################

"""
Contains ExcelConnector class and a help function.

:author: P. Vennström (VTT)
:date:   1.6.2019
"""

from itertools import islice, takewhile
import io
from openpyxl import load_workbook
from spinedb_api import (
    RelationshipClassMapping,
    ObjectClassMapping,
    ObjectGroupMapping,
    AlternativeMapping,
    ScenarioMapping,
    ScenarioAlternativeMapping,
    from_database,
    ParameterValueFormatError,
)
from ..io_api import SourceConnection


class ExcelConnector(SourceConnection):
    """Template class to read data from another QThread."""

    # name of data source, ex: "Text/CSV"
    DISPLAY_NAME = "Excel"

    # dict with option specification for source.
    OPTIONS = {
        "header": {"type": bool, "label": "Has header", "default": False},
        "row": {"type": int, "label": "Skip rows", "Minimum": 0, "default": 0},
        "column": {"type": int, "label": "Skip columns", "Minimum": 0, "default": 0},
        "read_until_col": {"type": bool, "label": "Read until empty column on first row", "default": False},
        "read_until_row": {"type": bool, "label": "Read until empty row on first column", "default": False},
    }

    FILE_EXTENSIONS = "*.xlsx;*.xlsm;*.xltx;*.xltm"

    def __init__(self, settings):
        super().__init__(settings)
        self._filename = None
        self._wb = None

    def connect_to_source(self, source):
        """saves filepath

        Arguments:
            source {str} -- filepath
        """
        if source:
            self._filename = source
            try:
                # FIXME: there seems to be no way of closing the workbook
                # when read_only=True, read file into memory first and then
                # open to avoid locking file while toolbox is running.
                with open(self._filename, "rb") as bin_file:
                    in_mem_file = io.BytesIO(bin_file.read())
                self._wb = load_workbook(in_mem_file, read_only=True)
            except Exception as error:
                raise error

    def disconnect(self):
        """Disconnect from connected source.
        """
        if self._wb:
            self._wb.close()
            self._wb = None
        self._filename = None

    def get_tables(self):
        """Method that should return Excel sheets as mappings and their options.

        Returns:
            dict: Sheets as mappings and options for each sheet or an empty dictionary if no workbook.

        Raises:
            Exception: If something goes wrong.
        """
        if not self._wb:
            return {}
        try:
            sheets = {}
            for sheet in self._wb.sheetnames:
                mapping, option = create_mapping_from_sheet(self._wb[sheet])
                sheets[sheet] = {"mapping": mapping, "options": option}
            return sheets
        except Exception as error:
            raise error

    def get_data_iterator(self, table, options, max_rows=-1):
        """
        Return data read from data source table in table. If max_rows is
        specified only that number of rows.
        """
        if not self._wb:
            return iter([]), [], 0

        if not table in self._wb:
            # table not found
            return iter([]), [], 0
        worksheet = self._wb[table]

        # get options
        has_header = options.get("header", False)
        skip_rows = options.get("row", 0)
        skip_columns = options.get("column", 0)
        stop_at_empty_col = options.get("read_until_col", False)
        stop_at_empty_row = options.get("read_until_row", False)

        if max_rows == -1:
            max_rows = None
        else:
            max_rows += skip_rows

        read_to_col = None
        try:
            first_row = next(islice(worksheet.iter_rows(), skip_rows, max_rows))
            if stop_at_empty_col:
                # find first empty col in top row and use that as a stop
                num_cols = 0
                for i, column in enumerate(islice(first_row, skip_columns, None)):

                    if column.value is None:
                        read_to_col = i + skip_columns
                        break
                    num_cols = num_cols + 1
            else:
                num_cols = len(first_row) - skip_columns
        except StopIteration:
            # no data
            num_cols = 0

        header = []
        rows = worksheet.iter_rows()
        rows = islice(rows, skip_rows, max_rows)
        # find header if it has one
        if has_header:
            try:
                header = [c.value for c in islice(next(rows), skip_columns, read_to_col)]
            except StopIteration:
                # no data
                return iter([]), [], 0

        # iterator for selected columns and and skipped rows
        data_iterator = (list(cell.value for cell in islice(row, skip_columns, read_to_col)) for row in rows)
        if stop_at_empty_row:
            # add condition to iterator
            condition = lambda row: row[0] is not None
            data_iterator = takewhile(condition, data_iterator)

        return data_iterator, header, num_cols

    def get_mapped_data(self, tables_mappings, options, table_types, table_row_types, max_rows=-1):
        """
        Overrides io_api method to check for some parameter_value types.
        """
        mapped_data, errors = super().get_mapped_data(tables_mappings, options, table_types, table_row_types, max_rows)
        for key in ("object_parameter_values", "relationship_parameter_values"):
            for index, value in enumerate(mapped_data.get(key, [])):
                val = value[-1]
                if isinstance(val, str) and val and val[0] == "{":
                    try:
                        val = from_database(val)
                        value = value[:-1] + (val,)
                        mapped_data[key][index] = value
                    except ParameterValueFormatError:
                        pass

        return mapped_data, errors


def get_mapped_data_from_xlsx(filepath):
    """Returns mapped data from given Excel file assuming it has the default Spine Excel format.

    Args:
        filepath (str): path to Excel file
    """
    connector = ExcelConnector(None)
    connector.connect_to_source(filepath)
    mappings_per_sheet = {}
    options_per_sheet = {}
    for sheet in connector._wb.sheetnames:
        mapping, options = create_mapping_from_sheet(connector._wb[sheet])
        if mapping is not None:
            mappings_per_sheet[sheet] = [mapping]
        if options is not None:
            options_per_sheet[sheet] = options
    types = row_types = dict.fromkeys(mappings_per_sheet, {})
    mapped_data, errors = connector.get_mapped_data(mappings_per_sheet, options_per_sheet, types, row_types)
    connector.disconnect()
    return mapped_data, errors


def create_mapping_from_sheet(worksheet):
    """
    Checks if sheet has the default Spine Excel format, if so creates a
    mapping object for each sheet.
    """

    options = {"header": False, "row": 0, "column": 0, "read_until_col": False, "read_until_row": False}
    mapping = ObjectClassMapping()
    sheet_type = worksheet["A2"].value
    sheet_data = worksheet["B2"].value
    if not isinstance(sheet_type, str):
        return None, None
    if not isinstance(sheet_data, str):
        return None, None
    if sheet_type.lower() not in [
        "relationship",
        "object",
        "object group",
        "alternative",
        "scenario",
        "scenario alternative",
    ]:
        return None, None
    if sheet_data.lower() not in ["parameter", "time series", "time pattern", "map", "array", "no data"]:
        return None, None
    if sheet_type.lower() == "relationship":
        mapping = RelationshipClassMapping()
        rel_dimension = worksheet["D2"].value
        rel_cls_name = worksheet["C2"].value
        if not isinstance(rel_cls_name, str):
            return None, None
        if not rel_cls_name:
            return None, None
        if not isinstance(rel_dimension, int):
            return None, None
        if rel_dimension < 1:
            return None, None
        if sheet_data.lower() == "parameter":
            obj_classes = next(islice(worksheet.iter_rows(), 3, 4))
            obj_classes = [r.value for r in obj_classes[:rel_dimension]]
        else:
            obj_classes = islice(worksheet.iter_rows(), 3, 3 + rel_dimension)
            obj_classes = [r[0].value for r in obj_classes]
        if not all(isinstance(r, str) for r in obj_classes) or any(r is None or r.isspace() for r in obj_classes):
            return None, None
        if sheet_data.lower() == "parameter":
            has_parameters = worksheet.cell(row=4, column=rel_dimension + 1).value is not None
            options.update({"header": True, "row": 3, "read_until_col": True, "read_until_row": True})
            map_dict = {
                "map_type": "RelationshipClass",
                "name": rel_cls_name,
                "object_classes": obj_classes,
                "objects": list(range(rel_dimension)),
            }
            if has_parameters:
                map_dict["parameters"] = {
                    "map_type": "parameter",
                    "name": {"map_type": "row", "value_reference": -1},
                    "parameter_type": "single value",
                }
            mapping = RelationshipClassMapping.from_dict(map_dict)
        elif sheet_data.lower() == "array":
            options.update({"header": False, "row": 3, "read_until_col": True, "read_until_row": False})
            mapping = RelationshipClassMapping.from_dict(
                {
                    "map_type": "RelationshipClass",
                    "name": rel_cls_name,
                    "skip_columns": [0],
                    "object_classes": obj_classes,
                    "objects": [{"map_type": "row", "value_reference": i} for i in range(rel_dimension)],
                    "parameters": {
                        "map_type": "parameter",
                        "name": {"map_type": "row", "value_reference": rel_dimension},
                        "extra_dimensions": [0],
                        "parameter_type": "array",
                    },
                }
            )
        elif sheet_data.lower() in ("time series", "time pattern"):
            options.update({"header": False, "row": 3, "read_until_col": True, "read_until_row": True})
            mapping = RelationshipClassMapping.from_dict(
                {
                    "map_type": "RelationshipClass",
                    "name": rel_cls_name,
                    "object_classes": obj_classes,
                    "objects": [{"map_type": "row", "value_reference": i} for i in range(rel_dimension)],
                    "parameters": {
                        "map_type": "parameter",
                        "name": {"map_type": "row", "value_reference": rel_dimension},
                        "extra_dimensions": [0],
                        "parameter_type": sheet_data.lower(),
                    },
                }
            )

    elif sheet_type.lower() == "object":
        obj_cls_name = worksheet["C2"].value
        if not isinstance(obj_cls_name, str):
            return None, None
        if not obj_cls_name:
            return None, None
        if sheet_data.lower() == "parameter":
            has_parameters = worksheet["B4"].value is not None
            options.update({"header": True, "row": 3, "read_until_col": True, "read_until_row": True})
            map_dict = {"map_type": "ObjectClass", "name": obj_cls_name, "objects": 0}
            if has_parameters:
                map_dict["parameters"] = {
                    "map_type": "parameter",
                    "name": {"map_type": "row", "value_reference": -1},
                    "parameter_type": "single value",
                }
            mapping = ObjectClassMapping.from_dict(map_dict)
        elif sheet_data.lower() == "array":
            options.update({"header": False, "row": 3, "read_until_col": True, "read_until_row": False})
            mapping = ObjectClassMapping.from_dict(
                {
                    "map_type": "ObjectClass",
                    "name": obj_cls_name,
                    "objects": {"map_type": "row", "value_reference": 0},
                    "skip_columns": [0],
                    "parameters": {
                        "map_type": "parameter",
                        "name": {"map_type": "row", "value_reference": 1},
                        "extra_dimensions": [0],
                        "parameter_type": "array",
                    },
                }
            )
        elif sheet_data.lower() in ("time series", "time pattern"):
            options.update({"header": False, "row": 3, "read_until_col": True, "read_until_row": True})
            mapping = ObjectClassMapping.from_dict(
                {
                    "map_type": "ObjectClass",
                    "name": obj_cls_name,
                    "objects": {"map_type": "row", "value_reference": 0},
                    "parameters": {
                        "map_type": "parameter",
                        "name": {"map_type": "row", "value_reference": 1},
                        "extra_dimensions": [0],
                        "parameter_type": sheet_data.lower(),
                    },
                }
            )

    elif sheet_type.lower() == "object group":
        obj_cls_name = worksheet["C2"].value
        if not isinstance(obj_cls_name, str):
            return None, None
        if not obj_cls_name:
            return None, None
        options.update({"header": True, "row": 3, "read_until_col": True, "read_until_row": True})
        mapping = ObjectGroupMapping.from_dict(
            {"map_type": "ObjectGroup", "name": obj_cls_name, "groups": 0, "members": 1}
        )
    elif sheet_type.lower() == "alternative":
        options.update({"header": True, "row": 3, "read_until_col": True, "read_until_row": True})
        mapping = AlternativeMapping.from_dict({"map_type": "Alternative", "name": 0})
    elif sheet_type.lower() == "scenario":
        options.update({"header": True, "row": 3, "read_until_col": True, "read_until_row": True})
        mapping = ScenarioMapping.from_dict({"map_type": "Scenario", "name": 0, "active": 1})
    elif sheet_type.lower() == "scenario alternative":
        options.update({"header": True, "row": 3, "read_until_col": True, "read_until_row": True})
        mapping = ScenarioAlternativeMapping.from_dict(
            {"map_type": "ScenarioAlternative", "scenario_name": 0, "alternative_name": 1, "before_alternative_name": 2}
        )
    else:
        return None, None
    return mapping, options
