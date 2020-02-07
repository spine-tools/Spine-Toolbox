######################################################################################################################
# Copyright (C) 2017-2020 Spine project consortium
# This file is part of Spine Toolbox.
# Spine Toolbox is free software: you can redistribute it and/or modify it under the terms of the GNU Lesser General
# Public License as published by the Free Software Foundation, either version 3 of the License, or (at your option)
# any later version. This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
# without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU Lesser General
# Public License for more details. You should have received a copy of the GNU Lesser General Public License along with
# this program. If not, see <http://www.gnu.org/licenses/>.
######################################################################################################################

"""
Framework for exporting a database to Excel file.

:author: P. Vennström (VTT), A. Soininen (VTT)
:date:   31.1.2020
"""

from itertools import groupby
from operator import itemgetter
import numpy as np
from openpyxl import Workbook
from spinedb_api import from_database, TimeSeries, TimePattern, DateTime, Duration, to_database


def _get_objects_and_parameters(db):
    """Exports all object data from spine database into unstacked list of lists

    Args:
        db (spinedb_api.DatabaseMapping): database mapping for database

    Returns:
        (List, List) First list contains parameter data, second one json data
    """

    # get all objects
    obj = db.object_list().all()

    # get all object classes
    obj_class = db.object_class_list().all()
    obj_class_id_2_name = {oc.id: oc.name for oc in obj_class}

    # get all parameter values
    pval = db.object_parameter_value_list().all()

    # get all parameter definitions
    par = db.object_parameter_definition_list().all()

    # make all in same format
    par = [(p.object_class_name, None, p.parameter_name, None) for p in par]
    pval = [(p.object_class_name, p.object_name, p.parameter_name, from_database(p.value)) for p in pval]
    obj = [(obj_class_id_2_name[p.class_id], p.name, None, None) for p in obj]
    obj_class = [(p.name, None, None, None) for p in obj_class]

    object_and_par = pval + par + obj + obj_class

    object_par = []
    object_json = []
    object_ts = []
    object_timepattern = []
    for d in object_and_par:
        if d[3] is None or isinstance(d[3], (int, float, str, DateTime, Duration)):
            object_par.append(d)
        elif isinstance(d[3], list):
            object_json.append(d)
            object_par.append(d[:-1] + (None,))
        elif isinstance(d[3], TimeSeries):
            object_ts.append(d)
            object_par.append(d[:-1] + (None,))
        elif isinstance(d[3], TimePattern):
            object_timepattern.append(d)
            object_par.append(d[:-1] + (None,))
        else:
            raise Warning(f"Unsuported export type: {type(d[3])}, Skipping export")

    return object_par, object_json, object_ts, object_timepattern


def _get_relationships_and_parameters(db):
    """Exports all relationship data from spine database into unstacked list of lists

    Args:
        db (spinedb_api.DatabaseMapping): database mapping for database

    Returns:
        (List, List) First list contains parameter data, second one json data
    """

    rel_class = db.wide_relationship_class_list().all()
    rel = db.wide_relationship_list().all()
    rel_par = db.relationship_parameter_definition_list().all()
    rel_par_value = db.relationship_parameter_value_list().all()

    rel_class_id_2_name = {rc.id: rc.name for rc in rel_class}

    out_data = [
        [r.relationship_class_name, r.object_name_list, r.parameter_name, from_database(r.value)] for r in rel_par_value
    ]

    rel_with_par = set(r.object_name_list for r in rel_par_value)
    rel_without_par = [
        [rel_class_id_2_name[r.class_id], r.object_name_list, None, None]
        for r in rel
        if r.object_name_list not in rel_with_par
    ]

    rel_class_par = [[r.relationship_class_name, None, r.parameter_name, None] for r in rel_par]

    rel_class_with_par = [r.relationship_class_name for r in rel_par]
    rel_class_without_par = [[r.name, None, None, None] for r in rel_class if r.name not in rel_class_with_par]

    rel_data = out_data + rel_without_par + rel_class_par + rel_class_without_par

    rel_par = []
    rel_json = []
    rel_ts = []
    rel_timepattern = []
    for d in rel_data:
        if d[3] is None or isinstance(d[3], (int, float, str, DateTime, Duration)):
            rel_par.append(d)
        elif isinstance(d[3], list):
            rel_json.append(d)
            rel_par.append(d[:-1] + [None])
        elif isinstance(d[3], TimeSeries):
            rel_ts.append(d)
            rel_par.append(d[:-1] + [None])
        elif isinstance(d[3], TimePattern):
            rel_timepattern.append(d)
            rel_par.append(d[:-1] + [None])
        else:
            raise Warning(f"Unsuported export type: {type(d[3])}, Skipping export")

    return rel_par, rel_json, rel_class, rel_ts, rel_timepattern


def _unstack_list_of_tuples(data, headers, key_cols, value_name_col, value_col):
    """Unstacks list of lists or list of tuples and creates a list of namedtuples
    whit unstacked data (pivoted data)

    Args:
        data (List[List]): List of lists with data to unstack
        headers (List[str]): List of header names for data
        key_cols (List[Int]): List of index for column that are keys, columns to not unstack
        value_name_col (Int): index to column containing name of data to unstack
        value_col (Int): index to column containing value to value_name_col

    Returns:
        (List[List]): List of list with headers in headers list
        (List): List of header names for each item in inner list
    """
    # find header names
    if isinstance(value_name_col, list) and len(value_name_col) > 1:
        value_name_getter = itemgetter(*value_name_col)
        value_names = sorted(
            set(value_name_getter(x) for x in data if not any(i is None for i in value_name_getter(x)))
        )
    else:
        if isinstance(value_name_col, list):
            value_name_col = value_name_col[0]
        value_name_getter = itemgetter(value_name_col)
        value_names = sorted(set(x[value_name_col] for x in data if x[value_name_col] is not None))

    key_names = [headers[n] for n in key_cols]
    # value_names = sorted(set(x[value_name_col] for x in data if x[value_name_col] is not None))
    headers = key_names + value_names

    # remove data with invalid key cols
    keyfunc = lambda x: [x[k] for k in key_cols]
    data = [x for x in data if None not in keyfunc(x)]
    data = sorted(data, key=keyfunc)

    # unstack data
    data_list_out = []
    for k, k_data in groupby(data, key=keyfunc):
        if None in k:
            continue
        line_data = [None] * len(value_names)
        for d in k_data:
            if value_name_getter(d) in value_names and d[value_col] is not None:
                line_data[value_names.index(value_name_getter(d))] = d[value_col]
        data_list_out.append(k + line_data)

    return data_list_out, headers


def _get_unstacked_relationships(db):
    """Gets all data for relationships in a unstacked list of list

    Args:
        db (spinedb_api.DatabaseMapping): database mapping for database

    Returns:
        (list, list, list, list): stacked relationships, stacked JSON, stacked time series and stacked time patterns
    """
    data, data_json, rel_class, data_ts, data_timepattern = _get_relationships_and_parameters(db)

    class_2_obj_list = {rc.name: rc.object_class_name_list.split(',') for rc in rel_class}

    keyfunc = lambda x: x[0]
    data_json = sorted(data_json, key=keyfunc)
    parsed_json = []
    # json data, split by relationship class
    for k, v in groupby(data_json, key=keyfunc):
        json_vals = []
        for row in v:
            rel_list = row[1].split(',')
            parameter = row[2]
            val = row[3]
            json_vals.append([rel_list + [parameter], val])
        if json_vals:
            object_classes = class_2_obj_list[k]
            parsed_json.append([k, object_classes, json_vals])

    data_ts = sorted(data_ts, key=keyfunc)
    parsed_ts = []
    # ts data, split by relationship class
    for k, v in groupby(data_ts, key=keyfunc):
        ts_vals = []
        for row in v:
            rel_list = row[1].split(',')
            parameter = row[2]
            val = row[3]
            ts_vals.append([rel_list + [parameter], val])
        if ts_vals:
            object_classes = class_2_obj_list[k]
            parsed_ts.append([k, object_classes, ts_vals])

    data_timepattern = sorted(data_timepattern, key=keyfunc)
    parsed_timepattern = []
    # ts data, split by relationship class
    for k, v in groupby(data_timepattern, key=keyfunc):
        tp_vals = []
        for row in v:
            rel_list = row[1].split(',')
            parameter = row[2]
            val = row[3]
            tp_vals.append([rel_list + [parameter], val])
        if tp_vals:
            object_classes = class_2_obj_list[k]
            parsed_timepattern.append([k, object_classes, tp_vals])

    # parameter data, split by relationship class
    stacked_rels = []
    data = sorted(data, key=keyfunc)
    for k, v in groupby(data, key=keyfunc):
        values = list(v)
        rel, par_names = _unstack_list_of_tuples(
            values, ["relationship_class", "relationship", "parameter", "value"], [0, 1], 2, 3
        )
        if rel:
            parameters = par_names[2:]
        else:
            parameters = list(set(p[2] for p in values))
        rel = [r[1].split(',') + list(r[2:]) for r in rel]
        object_classes = class_2_obj_list[k]
        stacked_rels.append([k, rel, object_classes, parameters])
    return stacked_rels, parsed_json, parsed_ts, parsed_timepattern


def _get_unstacked_objects(db):
    """Gets all data for objects in a unstacked list of list

    Args:
        db (spinedb_api.DatabaseMapping): database mapping for database

    Returns:
        (list, list, list, list): stacked objects, parsed JSON, parsed time series and parsed time patterns
    """
    data, data_json, data_ts, data_timepattern = _get_objects_and_parameters(db)

    keyfunc = lambda x: x[0]

    parsed_json = []
    data_json = sorted(data_json, key=keyfunc)
    for k, v in groupby(data_json, key=keyfunc):
        json_vals = []
        for row in v:
            obj = row[1]
            parameter = row[2]
            val = row[3]
            json_vals.append([[obj, parameter], val])
        if json_vals:
            parsed_json.append([k, [k], json_vals])

    parsed_ts = []
    data_ts = sorted(data_ts, key=keyfunc)
    for k, v in groupby(data_ts, key=keyfunc):
        ts_vals = []
        for row in v:
            obj = row[1]
            parameter = row[2]
            val = row[3]
            ts_vals.append([[obj, parameter], val])
        if ts_vals:
            parsed_ts.append([k, [k], ts_vals])

    data_timepattern = sorted(data_timepattern, key=keyfunc)
    parsed_timepattern = []
    # ts data, split by object class
    for k, v in groupby(data_timepattern, key=keyfunc):
        tp_vals = []
        for row in v:
            obj = row[1]
            parameter = row[2]
            val = row[3]
            tp_vals.append([[obj, parameter], val])
        if tp_vals:
            parsed_timepattern.append([k, [k], tp_vals])

    stacked_obj = []
    data = sorted(data, key=keyfunc)
    for k, v in groupby(data, key=keyfunc):
        values = list(v)
        obj, par_names = _unstack_list_of_tuples(values, ["object_class", "object", "parameter", "value"], [0, 1], 2, 3)
        if obj:
            parameters = par_names[2:]
        else:
            parameters = list(set(p[2] for p in values if p[2] is not None))
        obj = [[o[1]] + list(o[2:]) for o in obj]
        object_classes = [k]
        stacked_obj.append([k, obj, object_classes, parameters])
    return stacked_obj, parsed_json, parsed_ts, parsed_timepattern


def _write_relationships_to_xlsx(wb, relationship_data):
    """Writes Classes, parameter and parameter values for relationships.
    Writes one sheet per relationship class.

    Args:
        wb (openpyxl.Workbook): excel workbook to write too.
        relationship_data (List[List]): List of lists containing relationship
        data give by function get_unstacked_relationships
    """
    for rel in relationship_data:
        ws = wb.create_sheet()

        # try setting the sheet name to relationship class name
        # sheet name can only be 31 chars log
        title = "rel_" + rel[0]
        if len(title) < 32:
            ws.title = title

        ws['A1'] = "Sheet type"
        ws['A2'] = "relationship"
        ws['B1'] = "Data type"
        ws['B2'] = "Parameter"
        ws['C1'] = "relationship class name"
        ws['C2'] = rel[0]
        ws['D1'] = "Number of relationship dimensions"
        ws['D2'] = len(rel[2])
        ws['E1'] = "Number of pivoted relationship dimensions"
        ws['E2'] = 0

        for c, val in enumerate(rel[2]):
            ws.cell(row=4, column=c + 1).value = val

        for c, val in enumerate(rel[3]):
            ws.cell(row=4, column=len(rel[2]) + 1 + c).value = val

        start_row = 5
        start_col = 1
        for r, line in enumerate(rel[1]):
            for c, val in enumerate(line):
                if isinstance(val, (Duration, DateTime)):
                    val = to_database(val)
                ws.cell(row=start_row + r, column=start_col + c).value = val


def _write_json_array_to_xlsx(wb, data, sheet_type):
    """Writes json array data for object classes and relationship classes.
    Writes one sheet per relationship/object class.

    Args:
        wb (openpyxl.Workbook): excel workbook to write too.
        data (List[List]): List of lists containing json data give by function
        get_unstacked_objects and get_unstacked_relationships
        sheet_type (str): str with value "relationship" or "object" telling if data is for a relationship or object
    """
    for i, d in enumerate(data):
        if sheet_type == "relationship":
            sheet_title = "json_"
        elif sheet_type == "object":
            sheet_title = "json_"
        else:
            raise ValueError("sheet_type must be a str with value 'relationship' or 'object'")

        ws = wb.create_sheet()
        # sheet name can only be 31 chars log
        title = sheet_title + d[0]
        if len(title) < 32:
            ws.title = title
        else:
            ws.title = '{}_json{}'.format(sheet_type, i)

        ws['A1'] = "Sheet type"
        ws['A2'] = sheet_type
        ws['B1'] = "Data type"
        ws['B2'] = "1d array"
        ws['C1'] = sheet_type + " class name"
        ws['C2'] = d[0]

        if sheet_type == "relationship":
            ws['D1'] = "Number of relationship dimensions"
            ws['D2'] = len(d[1])

        title_rows = d[1] + ["json parameter"]
        for c, val in enumerate(title_rows):
            ws.cell(row=4 + c, column=1).value = val

        start_row = 4 + len(title_rows)
        for col, obj_list in enumerate(d[2]):
            for obj_iter, obj in enumerate(obj_list[0]):
                ws.cell(row=4 + obj_iter, column=2 + col).value = obj

            for row_iter, json_val in enumerate(obj_list[1]):
                ws.cell(row=start_row + row_iter, column=2 + col).value = json_val


def _write_TimeSeries_to_xlsx(wb, data, sheet_type, data_type):
    """Writes spinedb_api TimeSeries data for object classes and relationship classes.
    Writes one sheet per relationship/object class.

    Args:
        wb (openpyxl.Workbook): excel workbook to write too.
        data (List[List]): List of lists containing json data give by function
        get_unstacked_objects and get_unstacked_relationships
        sheet_type (str): str with value "relationship" or "object" telling if data is for a relationship or object
    """
    for i, d in enumerate(data):
        if sheet_type == "relationship":
            sheet_title = "ts_"
        elif sheet_type == "object":
            sheet_title = "ts_"
        else:
            raise ValueError("sheet_type must be a str with value 'relationship' or 'object'")

        if data_type.lower() == "time series":
            index_name = "timestamp"
        elif data_type.lower() == "time pattern":
            index_name = "pattern"
        else:
            raise ValueError("data_type must be a str with value 'time series' or 'time pattern'")

        ws = wb.create_sheet()
        # sheet name can only be 31 chars log
        title = sheet_title + d[0]
        if len(title) < 32:
            ws.title = title
        else:
            ws.title = '{}_ts{}'.format(sheet_type, i)

        ws['A1'] = "Sheet type"
        ws['A2'] = sheet_type
        ws['B1'] = "Data type"
        ws['B2'] = data_type
        ws['C1'] = sheet_type + " class name"
        ws['C2'] = d[0]

        if sheet_type == "relationship":
            ws['D1'] = "Number of relationship dimensions"
            ws['D2'] = len(d[1])

        title_rows = d[1] + [index_name]
        for c, val in enumerate(title_rows):
            ws.cell(row=4 + c, column=1).value = val

        # find common timestamps
        unique_timestamps = np.unique(np.concatenate([v[1].indexes for v in d[2]]))

        # write object names
        start_row = 4 + len(title_rows)
        for col, obj_list in enumerate(d[2]):
            for obj_iter, obj in enumerate(obj_list[0]):
                ws.cell(row=4 + obj_iter, column=2 + col).value = obj

        # write timestamps
        if data_type.lower() == "time series":
            for row_iter, time in enumerate(unique_timestamps):
                ws.cell(row=start_row + row_iter, column=1).value = str(np.datetime_as_string(time))
        else:
            for row_iter, time in enumerate(unique_timestamps):
                ws.cell(row=start_row + row_iter, column=1).value = str(time)

        # write values
        for col, obj_list in enumerate(d[2]):
            for row_index, value in zip(
                np.where(np.isin(unique_timestamps, obj_list[1].indexes))[0], obj_list[1].values
            ):
                ws.cell(row=start_row + row_index, column=2 + col).value = value


def _write_objects_to_xlsx(wb, object_data):
    """Writes Classes, parameter and parameter values for objects.
    Writes one sheet per relationship/object class.

    Args:
        wb (openpyxl.Workbook): excel workbook to write too.
        object_data (List[List]): List of lists containing relationship data give by function get_unstacked_objects
    """

    for i, obj in enumerate(object_data):
        ws = wb.create_sheet()

        # try setting the sheet name to object class name
        # sheet name can only be 31 chars log
        title = "obj_" + obj[0]
        if len(title) < 32:
            ws.title = title
        else:
            ws.title = "object_class{}".format(i)

        ws['A1'] = "Sheet type"
        ws['A2'] = "object"
        ws['B1'] = "Data type"
        ws['B2'] = "Parameter"
        ws['C1'] = "object class name"
        ws['C2'] = obj[0]

        for c, val in enumerate(obj[2]):
            ws.cell(row=4, column=c + 1).value = val

        for c, val in enumerate(obj[3]):
            ws.cell(row=4, column=len(obj[2]) + 1 + c).value = val

        start_row = 5
        start_col = 1
        for r, line in enumerate(obj[1]):
            for c, val in enumerate(line):
                if isinstance(val, (Duration, DateTime)):
                    val = to_database(val)
                ws.cell(row=start_row + r, column=start_col + c).value = val


def export_spine_database_to_xlsx(db, filepath):
    """Writes all data in a spine database into an excel file.

    Args:
        db (spinedb_api.DatabaseMapping): database mapping for database.
        filepath (str): str with filepath to save excel file to.
    """
    obj_data, obj_json_data, obj_ts, obj_timepattern = _get_unstacked_objects(db)
    rel_data, rel_json_data, rel_ts, rel_timepattern = _get_unstacked_relationships(db)
    wb = Workbook()
    _write_relationships_to_xlsx(wb, rel_data)
    _write_objects_to_xlsx(wb, obj_data)
    _write_json_array_to_xlsx(wb, obj_json_data, "object")
    _write_json_array_to_xlsx(wb, rel_json_data, "relationship")
    _write_TimeSeries_to_xlsx(wb, obj_ts, "object", "time series")
    _write_TimeSeries_to_xlsx(wb, rel_ts, "relationship", "time series")
    _write_TimeSeries_to_xlsx(wb, obj_timepattern, "object", "time pattern")
    _write_TimeSeries_to_xlsx(wb, rel_timepattern, "relationship", "time pattern")
    wb.save(filepath)
    wb.close()
